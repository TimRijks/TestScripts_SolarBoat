import serial
import openpyxl
from openpyxl import Workbook
import time
import os
import mysql.connector

PORT = "COM14"
BAUD = 9600
FILE = r"C:\Users\timri\OneDrive - HAN\Studiejaar2\S3_Project\Code\scripts\rs485_logAnT8.xlsx"

TAILSCALE_HOST = "100.75.92.26" # your MySQL server Tailscale IP
MYSQL_USER = "mse"
MYSQL_PASS = "embeddedBit"
MYSQL_DB   = "Ultrasoon"

db = mysql.connector.connect(
    host=TAILSCALE_HOST,
    user=MYSQL_USER,
    password=MYSQL_PASS,
    database=MYSQL_DB,
    port=3306
)
cursor = db.cursor()
def setup_excel():
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "RS485 Log"
        ws.append(["Timestamp", "data"])
        wb.save(FILE)

    wb = openpyxl.load_workbook(FILE)
    return wb, wb.active


def main():
    wb, ws = setup_excel()
    ser = serial.Serial(PORT, BAUD, timeout=1)
    
    buffer = b""  # Buffer om data te verzamelen
    start_time = time.time()

    try:
        while True:
            data = ser.read(64)

            if data:
                buffer += data
                
                # Zoek naar \r in de buffer
                while b'\r' in buffer:   
                    # Split bij de eerste \r
                    line, buffer = buffer.split(b'\r', 1)
                    
                    if line:  # Alleen opslaan als er data is
                        timestamp = time.time()
                        
                        ws.append([
                            timestamp,
                            line.decode('utf-8', errors='replace'),  # Decode naar string
                        ])
                        
                        print(timestamp, "\n", line, "\n")
                        
                        # Periodiek opslaan (elke 10 regels bijvoorbeeld)
                        if ws.max_row % 10 == 0:
                            wb.save(FILE)
                        
                        timestamp_ms = int((time.time() - start_time) * 1000)

                        sql = """
    INSERT INTO Analog_data
    (timestamp_ms, counts)
    VALUES (%s,%s)
    """ 
                        
                        try:
                            cursor.execute(sql, (timestamp_ms, line))
                            db.commit()
                            print("Inserted:", data)
                        except Exception as e:
                            print("DB error:", e)
                            continue

    except KeyboardInterrupt:
        # Sla eventuele resterende data op
        if buffer:
            timestamp =time.time()
            ws.append([timestamp, buffer.decode('utf-8', errors='replace')])
        
        print("Saving...")
        wb.save(FILE)
        ser.close()


if __name__ == "__main__":
    main()